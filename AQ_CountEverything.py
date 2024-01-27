import os
import arcpy
import openpyxl
import xlrd


def count():
    fc_list = arcpy.ListFeatureClasses()

    # this will not account for other excel file formats
    xlsx_files = arcpy.ListFiles('*.xlsx')

    xls_files = arcpy.ListFiles('*.xls')

    # counting the number of shapefiles and excel xlsx files in the folder
    count_shp = len(fc_list)
    count_exl = len(xlsx_files) + len(xls_files)
    print("(1)")
    print("File count:")
    print("A total of ", count_shp, " shapefiles and ",
          count_exl, " excel file(s) reside in this folder.")

    print()
    print()

    # counting number of features in shapefile
    print("(2)")
    print("Feature counts for shapefiles:")
    print()
    for i in fc_list:
        feature_count = arcpy.GetCount_management(i)
        print(
            f"Total number of features in {i}: {feature_count}")

    print()
    print()

    # counting rows in excel file(s):
    excel_files = xlsx_files + xls_files

    print("Row count for excel files:")
    print(
        f"The following is the list of excel files (.xlsx or .xls) in the workspace: {excel_files}")
    print()
    print()

    for file in excel_files:
        if file.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(file)
            sheets = workbook.sheetnames
            print(
                f"The following sheets' (within the {file} excel file) rows will be counted: {sheets}")
            print()
            for i in sheets:
                sht = workbook[i]
                num_rows = sht.max_row
                print(
                    f"    There are {num_rows} rows in {i} sheet of the {file} excel file. This includes any field name or other miscellaneous rows, should they exist.")
            print()
        elif file.endswith(".xls"):
            workbook_xls = xlrd.open_workbook(file)
            sheets_xls = [sheet.name for sheet in workbook_xls.sheets()]
            print(
                f"The following sheets' (within the {file} excel file) rows will be counted: {sheets_xls}")
            print()
            for i in sheets_xls:
                sht_xls = workbook_xls.sheet_by_name(i)
                rows_xls = sht_xls.nrows
                print(
                    f"    There are {rows_xls} rows in {i} sheet of the {file} excel file. This includes any field name or other miscellaneous rows, should they exist.")
    print()
    print()

# listing field names for point shapefiles


def pointlist():
    fc_pointlist = arcpy.ListFeatureClasses(feature_type='Point')

    print()
    print("(3)")
    print("Field list(s) for point feature classes:")
    print(
        f"The following is the list of point feature classes in the workspace: {fc_pointlist}")
    print()

    for i in fc_pointlist:
        fc_pointfields = arcpy.ListFields(
            i)
        print(f"The point feature class {i} has fields: ")
        print()
        for field in fc_pointfields:
            print(f"{field.name}")
        print()


def main():
    while True:
        print()
        print("Hello.")
        print()
        print("This program will: (1) count the numbers of excel and shapefiles in a specified workspace folder, (2) count the total numbers of rows and/or features for the excel files and/or shapefiles respectively, and (3) list and report the fields for point shapefiles. Please be advised that this program will only account for excel files with the extension .xlsx or .xls. ")
        print()

        while True:
            work_direct = input(
                "Enter the path to the folder with the files you would like counted: ")
            print()
            if os.path.exists(work_direct) and os.path.isdir(work_direct):
                os.chdir(work_direct)
                arcpy.env.workspace = work_direct
                print("The working directory is now set to: ", os.getcwd())
                print()
                break
            else:
                print("invalid working directory path. Try again.")
                print()

        count()
        pointlist()
        rerun_option = input(
            "If you are finished with this program, please type Y, otherwise if you would like to use the program again, type N: ")
        if rerun_option == "Y" or rerun_option == "y" or rerun_option == "YES" or rerun_option == "yes" or rerun_option == "Yes":
            print()
            print("End of program. Thank you.\nGoodbye.")
            break
        if rerun_option == "N" or rerun_option == "n" or rerun_option == "NO" or rerun_option == "no" or rerun_option == "No":
            continue
        else:
            print("Invalid response. Please rerun the program.\nGoodbye.")
            break


if __name__ == "__main__":
    main()
